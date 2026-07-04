"""XLSX → ParsedDocument converter using pandas + openpyxl."""

from __future__ import annotations

import html
from math import isnan
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from doc_converter.converters.base import BaseConverter
from doc_converter.ir import DocElement, ParsedDocument


def _escape_md_cell(value: str) -> str:
    """Escape pipe characters for markdown table cells."""
    return value.replace("|", "\\|").replace("\n", " ")


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and isnan(value):
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _build_merged_lookup(sheet: Worksheet) -> dict[tuple[int, int], tuple[int, int, str]]:
    """Map each cell coordinate to its merge anchor and shared value."""
    lookup: dict[tuple[int, int], tuple[int, int, str]] = {}
    for merge_range in sheet.merged_cells.ranges:
        min_row, min_col = merge_range.min_row, merge_range.min_col
        anchor_value = _cell_to_str(sheet.cell(min_row, min_col).value)
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                lookup[(row, col)] = (min_row, min_col, anchor_value)
    return lookup


def _sheet_dimensions(sheet: Worksheet) -> tuple[int, int]:
    max_row = sheet.max_row or 0
    max_col = sheet.max_column or 0
    return max_row, max_col


def _collect_grid(sheet: Worksheet, df: pd.DataFrame) -> list[list[str]]:
    """Build a 2D grid honoring merged cells from openpyxl."""
    max_row, max_col = _sheet_dimensions(sheet)
    if max_row == 0 or max_col == 0:
        return []

    merged = _build_merged_lookup(sheet)
    grid: list[list[str]] = []

    for row_idx in range(1, max_row + 1):
        row_values: list[str] = []
        for col_idx in range(1, max_col + 1):
            if (row_idx, col_idx) in merged:
                _, _, value = merged[(row_idx, col_idx)]
            else:
                df_row = row_idx - 1
                df_col = col_idx - 1
                if df_row < len(df.index) and df_col < len(df.columns):
                    value = _cell_to_str(df.iat[df_row, df_col])
                else:
                    value = _cell_to_str(sheet.cell(row_idx, col_idx).value)
            row_values.append(value)
        grid.append(row_values)

    return grid


def _trim_empty_edges(grid: list[list[str]]) -> list[list[str]]:
    if not grid:
        return grid

    def row_empty(row: list[str]) -> bool:
        return all(not cell for cell in row)

    while grid and row_empty(grid[-1]):
        grid.pop()
    while grid and all(not row[0] for row in grid if row):
        for row in grid:
            if row:
                row.pop(0)
    while grid and all(not row[-1] for row in grid if row):
        for row in grid:
            if row:
                row.pop()
    return grid


def _grid_to_markdown(grid: list[list[str]]) -> str:
    if not grid:
        return ""

    col_count = max(len(row) for row in grid)
    normalized = [row + [""] * (col_count - len(row)) for row in grid]

    header = normalized[0]
    body = normalized[1:] if len(normalized) > 1 else [[""] * col_count]

    lines = [
        "| " + " | ".join(_escape_md_cell(c) for c in header) + " |",
        "| " + " | ".join("---" for _ in header) + " |",
    ]
    for row in body:
        lines.append("| " + " | ".join(_escape_md_cell(c) for c in row) + " |")
    return "\n".join(lines)


def _merge_span_at(
    sheet: Worksheet,
    row: int,
    col: int,
) -> tuple[int, int] | None:
    """Return (rowspan, colspan) if (row, col) is the top-left of a merge."""
    for merge_range in sheet.merged_cells.ranges:
        if merge_range.min_row == row and merge_range.min_col == col:
            rowspan = merge_range.max_row - merge_range.min_row + 1
            colspan = merge_range.max_col - merge_range.min_col + 1
            if rowspan > 1 or colspan > 1:
                return rowspan, colspan
    return None


def _is_covered_by_merge(sheet: Worksheet, row: int, col: int) -> bool:
    for merge_range in sheet.merged_cells.ranges:
        if (
            merge_range.min_row <= row <= merge_range.max_row
            and merge_range.min_col <= col <= merge_range.max_col
            and not (row == merge_range.min_row and col == merge_range.min_col)
        ):
            return True
    return False


def _grid_to_html(sheet: Worksheet, grid: list[list[str]]) -> str:
    if not grid:
        return "<table></table>"

    rows_html: list[str] = []
    for row_idx, row in enumerate(grid, start=1):
        cells: list[str] = []
        for col_idx, value in enumerate(row, start=1):
            if _is_covered_by_merge(sheet, row_idx, col_idx):
                continue
            span = _merge_span_at(sheet, row_idx, col_idx)
            attrs = ""
            if span:
                rowspan, colspan = span
                attrs = f' rowspan="{rowspan}" colspan="{colspan}"'
            cells.append(f"<td{attrs}>{html.escape(value)}</td>")
        rows_html.append(f"<tr>{''.join(cells)}</tr>")

    return (
        '<table style="border-collapse:collapse;width:100%;">'
        '<tbody>'
        f'{"".join(rows_html)}'
        "</tbody></table>"
    )


class XlsxConverter(BaseConverter):
    """Convert Excel workbooks into heading + table elements per sheet."""

    def parse(self, path: Path) -> ParsedDocument:
        path = path.resolve()
        workbook = load_workbook(path, data_only=True)
        sheets_data: dict[str, pd.DataFrame] = pd.read_excel(path, sheet_name=None, header=None)

        elements: list[DocElement] = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            df = sheets_data.get(sheet_name, pd.DataFrame())
            grid = _trim_empty_edges(_collect_grid(sheet, df))
            if not grid:
                continue

            elements.append(
                DocElement(
                    type="heading",
                    content=sheet_name,
                    level=2,
                    source_sheet=sheet_name,
                    extraction_method="openpyxl+pandas",
                )
            )
            elements.append(
                DocElement(
                    type="table",
                    content=_grid_to_markdown(grid),
                    html_content=_grid_to_html(sheet, grid),
                    source_sheet=sheet_name,
                    extraction_method="openpyxl+pandas",
                    confidence=1.0,
                )
            )

        return ParsedDocument(
            source_file=str(path),
            source_type="xlsx",
            elements=elements,
        )
